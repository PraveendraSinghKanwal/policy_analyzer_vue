import os
import json
from openpyxl import load_workbook
import glob
import re

# Constants for better maintainability
PIXELS_PER_EXCEL_UNIT = 7
PIXELS_PER_CHARACTER = 10
MIN_COLUMN_WIDTH = 100
MAX_COLUMN_WIDTH = 500
DEFAULT_ROW_HEIGHT = 20
DEFAULT_COLUMN_WIDTH = 100
POINTS_TO_PIXELS_RATIO = 0.75

def extract_cell_style(cell):
    """
    Extract styling information from a cell and return as array
    Returns: [cell_value, color, bg_color, fontWeight, fontStyle, textDecoration, fontSize, fontFamily, textAlign, verticalAlign, columnWidth, rowHeight]
    """
    # Helper function to safely convert RGB to hex
    def rgb_to_hex(rgb_value):
        if rgb_value is None:
            return "#FFFFFF"
        
        try:
            # Convert to string
            rgb_str = str(rgb_value)
            
            # Handle different RGB formats
            if rgb_str.startswith('FF'):
                # Remove 'FF' prefix (alpha channel)
                return "#" + rgb_str[2:]
            elif rgb_str.startswith('00'):
                # Handle cases like "00000000" or "00FFFFFF"
                # Remove the first two characters (alpha channel)
                return "#" + rgb_str[2:]
            elif rgb_str.startswith('#'):
                return rgb_str
            elif len(rgb_str) == 8:
                # 8-character RGB with alpha channel (e.g., "00000000")
                return "#" + rgb_str[2:]  # Remove first 2 chars (alpha)
            elif len(rgb_str) == 6:
                # 6-character RGB without alpha
                return "#" + rgb_str
            else:
                # Default case
                return "#" + rgb_str
        except Exception:
            return "#FFFFFF"
    
    # Helper function to determine theme color from context when direct access fails
    def determine_theme_color_from_context(cell):
        """
        Try to determine the correct theme color when direct theme access fails.
        This function uses heuristics based on cell context, content, and surrounding cells.
        """
        # Get cell properties
        row = cell.row
        column = cell.column
        cell_value = str(cell.value).lower() if cell.value else ""
        
        # Check if this is likely a header cell (first row)
        if row == 1:
            return "#0070C0"  # Blue for headers
        
        # Check if this is likely a category cell (column B)
        if column == 2:
            return "#D6DCE4"  # Light gray for categories
        
        # Check if this is likely a subcategory cell (column C)
        if column == 3:
            # Use intelligent color detection based on content
            return get_intelligent_color(cell)
        
        # Check if this is likely a content cell (column D)
        if column == 4:
            # Use intelligent color detection based on content
            return get_intelligent_color(cell)
        
        # For other columns, use position-based logic
        if column == 1:  # Usually row numbers
            return "#FFFFFF"  # White for row numbers
        
        # Default fallback
        return "#FFFFFF"
    
    # Helper function to get a more intelligent color based on surrounding context
    def get_intelligent_color(cell):
        """
        Get a more intelligent color by analyzing the cell's context and content.
        This function tries to mimic Excel's conditional formatting logic.
        """
        cell_value = str(cell.value).lower() if cell.value else ""
        
        # Keywords that might indicate special colors
        priority_keywords = ["preferred", "primary", "recommended", "best"]
        warning_keywords = ["prohibited", "forbidden", "not allowed", "restricted"]
        info_keywords = ["optional", "suggested", "alternative"]
        action_keywords = ["required", "mandatory", "must", "shall"]
        
        # Check for priority content (usually highlighted)
        if any(keyword in cell_value for keyword in priority_keywords):
            return "#FFFF00"  # Yellow for priority items
        
        # Check for warning content (usually red or orange)
        if any(keyword in cell_value for keyword in warning_keywords):
            return "#FF6B6B"  # Light red for warnings
        
        # Check for informational content (usually blue or gray)
        if any(keyword in cell_value for keyword in info_keywords):
            return "#E3F2FD"  # Light blue for info
        
        # Check for action content (usually green or highlighted)
        if any(keyword in cell_value for keyword in action_keywords):
            return "#C8E6C9"  # Light green for actions
        
        # Default to white for regular content
        return "#FFFFFF"
    
    # Helper function to handle theme colors
    def theme_to_hex(theme_value):
        # Excel theme color mapping (basic mapping)
        theme_colors = {
            0: "#FFFFFF",  # White
            1: "#000000",  # Black
            2: "#E7E6E6",  # Light Gray
            3: "#44546A",  # Dark Gray
            4: "#4472C4",  # Blue
            5: "#ED7D31",  # Orange
            6: "#A5A5A5",  # Gray
            7: "#FFC000",  # Yellow
            8: "#5B9BD5",  # Light Blue
            9: "#70AD47",  # Green
            10: "#FF0000", # Red
            11: "#7030A0", # Purple
        }
        
        # Handle case where theme_value might be a string error
        if isinstance(theme_value, str):
            # Check if it's the specific error message we're seeing
            if "Values must be of type" in theme_value:
                # Return None to indicate we need context-based determination
                return None
            return "#FFFFFF"  # Default to white for other string errors
        
        return theme_colors.get(theme_value, "#FFFFFF")
    
    # Helper function to extract background color from fill
    def extract_background_color(fill):
        """Extract background color from cell fill, handling all pattern types"""
        if not fill:
            return "#FFFFFF"
        
        # Check for pattern type
        pattern_type = getattr(fill, 'patternType', None)
        
        if pattern_type == 'solid':
            # Solid fill - check start_color
            if fill.start_color:
                if fill.start_color.rgb:
                    rgb_str = str(fill.start_color.rgb)
                    # Handle special case where RGB is 00000000 (transparent/white)
                    if rgb_str == '00000000':
                        return "#FFFFFF"
                    else:
                        return rgb_to_hex(fill.start_color.rgb)
                elif hasattr(fill.start_color, 'theme') and fill.start_color.theme is not None:
                    try:
                        theme_value = fill.start_color.theme
                        theme_color = theme_to_hex(theme_value)
                        if theme_color is None:
                            # theme_to_hex returned None, use context-based determination
                            return determine_theme_color_from_context(cell)
                        else:
                            return theme_color
                    except Exception:
                        return "#FFFFFF"  # Default to white if theme access fails
                elif hasattr(fill.start_color, 'type') and fill.start_color.type == 'indexed':
                    # Handle indexed colors (Excel built-in palette)
                    return "#FFFFFF"  # Default for indexed colors
        elif pattern_type == 'none' or pattern_type is None:
            # No fill pattern - should be transparent/white
            return "#FFFFFF"
        else:
            # Other pattern types - check start_color as fallback
            if fill.start_color:
                if fill.start_color.rgb:
                    rgb_str = str(fill.start_color.rgb)
                    # Handle special case where RGB is 00000000 (transparent/white)
                    if rgb_str == '00000000':
                        return "#FFFFFF"
                    else:
                        return rgb_to_hex(fill.start_color.rgb)
                elif hasattr(fill.start_color, 'theme') and fill.start_color.theme is not None:
                    try:
                        theme_value = fill.start_color.theme
                        theme_color = theme_to_hex(theme_value)
                        if theme_color is None:
                            # theme_to_hex returned None, use context-based determination
                            return determine_theme_color_from_context(cell)
                        else:
                            return theme_color
                    except Exception:
                        return "#FFFFFF"  # Default to white if theme access fails
        
        return "#FFFFFF"
    
    # Get cell value
    cell_value = str(cell.value) if cell.value is not None else ""
    
    # Default values
    color = "#000000"
    bg_color = "#FFFFFF"  # Default white background
    font_weight = "normal"
    font_style = "normal"
    text_decoration = "none"
    font_size = "11pt"
    font_family = "Calibri"
    text_align = "left"
    vertical_align = "top"
    column_width = DEFAULT_COLUMN_WIDTH
    row_height = DEFAULT_ROW_HEIGHT
    
    # Extract font properties
    if cell.font:
        # Font color
        if cell.font.color:
            if cell.font.color.rgb:
                color = rgb_to_hex(cell.font.color.rgb)
            elif hasattr(cell.font.color, 'theme') and cell.font.color.theme is not None:
                try:
                    theme_color = theme_to_hex(cell.font.color.theme)
                    if theme_color is None:
                        # theme_to_hex returned None, use default black
                        color = "#000000"
                    else:
                        color = theme_color
                except Exception:
                    color = "#000000"  # Default to black if theme access fails
        
        # Font weight
        if cell.font.bold:
            font_weight = "bold"
        
        # Font style
        if cell.font.italic:
            font_style = "italic"
        
        # Text decoration
        if cell.font.underline:
            text_decoration = "underline"
        
        # Font size
        if cell.font.size:
            font_size = f"{int(cell.font.size)}pt"
        
        # Font family
        if cell.font.name:
            font_family = cell.font.name
    
    # Extract background color using the helper function
    bg_color = extract_background_color(cell.fill)
    
    # Extract alignment
    if cell.alignment:
        if cell.alignment.horizontal:
            text_align = cell.alignment.horizontal
        if cell.alignment.vertical:
            vertical_align = cell.alignment.vertical
    
    return [
        cell_value,
        color,
        bg_color,
        font_weight,
        font_style,
        text_decoration,
        font_size,
        font_family,
        text_align,
        vertical_align,
        column_width,
        row_height
    ]

def process_excel_file(file_path):
    """
    Process a single Excel file and return its data in the required format
    """
    try:
        # Load the workbook
        workbook = load_workbook(file_path, data_only=True)
        
        # Get the first worksheet
        worksheet = workbook.active
        
        # Get file name
        file_name = os.path.basename(file_path)
        
        # Get the used range
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Initialize cell_data array
        cell_data = []
        
        # Calculate column widths once from the header row (first row)
        column_widths = []
        for col_idx in range(1, max_col + 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            column_dimension = worksheet.column_dimensions[col_letter]
            
            # Get the column width
            if column_dimension.width is not None:
                # Use the explicit width from Excel
                column_widths.append(int(column_dimension.width * PIXELS_PER_EXCEL_UNIT))
            else:
                # Calculate width based on content or use default
                # Try to estimate width from the header cell content and column content
                header_cell = worksheet.cell(row=1, column=col_idx)
                header_value = str(header_cell.value) if header_cell.value else ""
                
                # Also check some content cells in this column to get better estimate
                max_content_length = len(header_value)
                for row_idx in range(2, min(max_row + 1, 10)):  # Check first 9 data rows
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = str(cell.value) if cell.value else ""
                    max_content_length = max(max_content_length, len(cell_value))
                
                # Estimate width based on content length (improved calculation)
                # Use PIXELS_PER_CHARACTER per character for better readability, minimum MIN_COLUMN_WIDTH
                estimated_width = max(max_content_length * PIXELS_PER_CHARACTER, MIN_COLUMN_WIDTH)
                
                # Cap at reasonable maximum (MAX_COLUMN_WIDTH)
                estimated_width = min(estimated_width, MAX_COLUMN_WIDTH)
                
                column_widths.append(estimated_width)
        
        # Process each row
        for row_idx in range(1, max_row + 1):
            row_data = []
            
            # Process each column in the row
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell_style = extract_cell_style(cell)
                
                # Apply the column width calculated from header row
                cell_style[10] = column_widths[col_idx - 1]
                
                # Get row height for this specific row
                row_height = worksheet.row_dimensions[row_idx].height
                if row_height:
                    cell_style[11] = int(row_height * POINTS_TO_PIXELS_RATIO)
                
                row_data.append(cell_style)
            
            cell_data.append(row_data)
        
        return {
            "name": file_name,
            "data": cell_data
        }
        
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return None

def sanitize_filename(filename):
    """
    Convert filename to a safe JSON filename
    """
    # Remove file extension
    name_without_ext = os.path.splitext(filename)[0]
    
    # Replace special characters with underscores
    safe_name = re.sub(r'[^a-zA-Z0-9_&-]', '_', name_without_ext)
    
    # Remove multiple consecutive underscores
    safe_name = re.sub(r'_+', '_', safe_name)
    
    # Remove leading/trailing underscores
    safe_name = safe_name.strip('_')
    
    return safe_name

def create_individual_json_files(input_folder, output_folder):
    """
    Read all Excel files from input folder and create individual JSON files in output folder
    """
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created output folder: {output_folder}")
    
    # Supported Excel file extensions
    excel_extensions = ['*.xlsx', '*.xls']
    
    # Find all Excel files in the input folder
    excel_files = []
    for extension in excel_extensions:
        excel_files.extend(glob.glob(os.path.join(input_folder, extension)))
    
    if not excel_files:
        print(f"No Excel files found in {input_folder}")
        return
    
    print(f"Found {len(excel_files)} Excel files:")
    for file in excel_files:
        print(f"  - {os.path.basename(file)}")
    
    # Process each Excel file and create individual JSON
    processed_count = 0
    failed_count = 0
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        print(f"\nProcessing: {file_name}")
        
        # Process the Excel file
        file_data = process_excel_file(file_path)
    
        if file_data:
            # Create safe filename for JSON
            safe_name = sanitize_filename(file_name)
            json_filename = f"{safe_name}.json"
            json_filepath = os.path.join(output_folder, json_filename)
            
            # Write individual JSON file
            try:
                with open(json_filepath, 'w', encoding='utf-8') as f:
                    json.dump(file_data, f, indent=2, ensure_ascii=False)
                
                print(f"  ✓ Created: {json_filename}")
                print(f"    Rows: {len(file_data['data'])}")
                if file_data['data']:
                    print(f"    Columns: {len(file_data['data'][0])}")
                
                processed_count += 1
                
            except Exception as e:
                print(f"  ✗ Error writing {json_filename}: {str(e)}")
                failed_count += 1
        else:
            print(f"  ✗ Failed to process {file_name}")
            failed_count += 1
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"PROCESSING SUMMARY:")
    print(f"  Total files found: {len(excel_files)}")
    print(f"  Successfully processed: {processed_count}")
    print(f"  Failed: {failed_count}")
    print(f"  Output folder: {output_folder}")
    
    if processed_count > 0:
        print(f"\nJSON files created:")
        for file in os.listdir(output_folder):
            if file.endswith('.json'):
                file_path = os.path.join(output_folder, file)
                file_size = os.path.getsize(file_path)
                print(f"  - {file} ({file_size:,} bytes)")

def process_excel_files():
    """
    Main function to process all Excel files
    """
    # Configuration
    input_folder = "Analysis"  # Folder containing Excel files
    output_folder = "excel_data"  # Folder where JSON files will be saved
    
    # Create individual JSON files for each Excel file
    create_individual_json_files(input_folder, output_folder)
    
    print(f"\n✓ Processing complete!")